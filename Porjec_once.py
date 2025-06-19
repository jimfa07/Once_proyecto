import streamlit as st
from datetime import datetime, date, timedelta
from io import BytesIO
import os
import csv
import pickle
import openpyxl
from openpyxl import Workbook
from supabase import create_client, Client
import json

# --- CONFIGURACIÓN INICIAL Y CONSTANTES ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# Configuración de Supabase
SUPABASE_URL = "https://crmrihkokjnvhzqywldv.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNybXJpaGtva2pudmh6cXl3bGR2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTAyMzAyNTEsImV4cCI6MjA2NTgwNjI1MX0.ZM0xeEiWikvyeOGzNPrpMouNyaoAU_w225acXF6R1f8"

# Cliente de Supabase
@st.cache_resource
def init_supabase():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# Archivos de datos del sistema principal (proveedores)
DATA_FILE = os.path.join(DATA_DIR, "registro_data.csv")
DEPOSITS_FILE = os.path.join(DATA_DIR, "registro_depositos.csv")
DEBIT_NOTES_FILE = os.path.join(DATA_DIR, "registro_notas_debito.csv")

# Archivos de datos del sistema secundario (ventas/gastos)
VENTAS_FILE = os.path.join(DATA_DIR, 'ventas.csv')
GASTOS_FILE = os.path.join(DATA_DIR, 'gastos.csv')

# Constantes del sistema principal
INITIAL_ACCUMULATED_BALANCE = 44.64 
PRODUCT_NAME = "Pollo"
LBS_PER_KG = 2.20462

PROVEEDORES = ["LIRIS SA", "Gallina 1", "Monze Anzules", "Medina"] 
TIPOS_DOCUMENTO = ["Factura", "Nota de debito", "Nota de credito"]
AGENCIAS = [
    "Cajero Automatico Pichincha", "Cajero Automatico Pacifico",
    "Cajero Automatico Guayaquil", "Cajero Automatico Bolivariano",
    "Banco Pichincha", "Banco del Pacifico", "Banco de Guayaquil",
    "Banco Bolivariano"
]

# Columnas para el sistema principal
COLUMNS_DATA = [
    "N", "Fecha", "Proveedor", "Producto", "Cantidad",
    "Peso Salida (kg)", "Peso Entrada (kg)", "Tipo Documento",
    "Cantidad de gavetas", "Precio Unitario ($)", "Promedio",
    "Kilos Restantes", "Libras Restantes", "Total ($)",
    "Monto Deposito", "Saldo diario", "Saldo Acumulado"
]
COLUMNS_DEPOSITS = ["Fecha", "Empresa", "Agencia", "Monto", "Documento", "N"]
COLUMNS_DEBIT_NOTES = ["Fecha", "Libras calculadas", "Descuento", "Descuento posible", "Descuento real"]

# Constantes del sistema secundario
CLIENTES = [
    "D. Vicente", "D. Jorge", "D. Quinde", "Sra. Isabel", "Sra. Alba",
    "Sra Yolanda", "Sra Laura Mercado", "D. Segundo", "Legumbrero",
    "Peruana Posorja", "Sra. Sofia", "Sra. Jessica", "Sra Alado de Jessica",
    "Comedor Gordo Posorja", "Patitas Posorja", "Sra. Celeste", "Caro negro", "Tienda Isabel Posorja",
    "Carnicero Posorja", "Moreira","Senel", "Chuzos Narcisa", "Eddy", "D. Jonny", "D. Sra Madelyn", "Lobo Mercado"
]

TIPOS_AVE = ["Pollo", "Gallina"]

CATEGORIAS_GASTO = [
    "G. Alimentación", "G. Transporte", "G. Producción", "G. Salud",
    "G. Educación", "G. Mano de obra", "G. Pérdida", "G. Varios", "Otros Gastos"
]

# Configuración de la página de Streamlit
st.set_page_config(
    page_title="Sistema Integral de Gestión - Proveedores y Ventas", 
    layout="wide", 
    initial_sidebar_state="expanded"
)

# --- FUNCIONES AUXILIARES ---

def formatear_moneda(valor):
    """Formatea un valor numérico como una cadena de moneda."""
    try:
        return f"${float(valor):,.2f}"
    except:
        return "$0.00"

def parse_float(value, default=0.0):
    """Convierte un valor a float de forma segura."""
    try:
        return float(value) if value != '' and value is not None else default
    except (ValueError, TypeError):
        return default

def parse_int(value, default=0):
    """Convierte un valor a int de forma segura."""
    try:
        return int(float(value)) if value != '' and value is not None else default
    except (ValueError, TypeError):
        return default

def parse_date(value):
    """Convierte un valor a date de forma segura."""
    try:
        if isinstance(value, str):
            return datetime.strptime(value, '%Y-%m-%d').date()
        elif isinstance(value, date):
            return value
        elif hasattr(value, 'date'):
            return value.date()
        else:
            return datetime.now().date()
    except:
        return datetime.now().date()

# --- FUNCIONES DE SUPABASE ---

def sync_to_supabase():
    """Sincroniza todos los datos locales con Supabase."""
    try:
        supabase = init_supabase()
        
        # Sincronizar registros principales
        if st.session_state.data_records:
            for record in st.session_state.data_records:
                record_data = record.copy()
                # Convertir todos los valores a tipos apropiados para Supabase
                for key, value in record_data.items():
                    if value == '':
                        record_data[key] = None
                
                # Intentar insertar o actualizar
                result = supabase.table('registros_proveedores').upsert(record_data).execute()
        
        # Sincronizar depósitos
        if st.session_state.deposit_records:
            for record in st.session_state.deposit_records:
                record_data = record.copy()
                for key, value in record_data.items():
                    if value == '':
                        record_data[key] = None
                supabase.table('depositos').upsert(record_data).execute()
        
        # Sincronizar notas de débito
        if st.session_state.debit_records:
            for record in st.session_state.debit_records:
                record_data = record.copy()
                for key, value in record_data.items():
                    if value == '':
                        record_data[key] = None
                supabase.table('notas_debito').upsert(record_data).execute()
        
        # Sincronizar ventas
        if st.session_state.ventas_records:
            for record in st.session_state.ventas_records:
                record_data = record.copy()
                for key, value in record_data.items():
                    if value == '':
                        record_data[key] = None
                supabase.table('ventas').upsert(record_data).execute()
        
        # Sincronizar gastos
        if st.session_state.gastos_records:
            for record in st.session_state.gastos_records:
                record_data = record.copy()
                for key, value in record_data.items():
                    if value == '':
                        record_data[key] = None
                supabase.table('gastos').upsert(record_data).execute()
        
        return True
    except Exception as e:
        st.error(f"Error sincronizando con Supabase: {e}")
        return False

def load_from_supabase():
    """Carga todos los datos desde Supabase."""
    try:
        supabase = init_supabase()
        
        # Cargar registros principales
        try:
            result = supabase.table('registros_proveedores').select('*').execute()
            if result.data:
                st.session_state.data_records = result.data
        except:
            pass  # Tabla puede no existir aún
        
        # Cargar depósitos
        try:
            result = supabase.table('depositos').select('*').execute()
            if result.data:
                st.session_state.deposit_records = result.data
        except:
            pass
        
        # Cargar notas de débito
        try:
            result = supabase.table('notas_debito').select('*').execute()
            if result.data:
                st.session_state.debit_records = result.data
        except:
            pass
        
        # Cargar ventas
        try:
            result = supabase.table('ventas').select('*').execute()
            if result.data:
                st.session_state.ventas_records = result.data
        except:
            pass
        
        # Cargar gastos
        try:
            result = supabase.table('gastos').select('*').execute()
            if result.data:
                st.session_state.gastos_records = result.data
        except:
            pass
        
        return True
    except Exception as e:
        st.error(f"Error cargando desde Supabase: {e}")
        return False

# --- FUNCIONES DE EXCEL ---

def create_excel_export():
    """Crea un archivo Excel con todos los datos del sistema."""
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # Hoja de Registros Principales
    ws_registros = wb.create_sheet("Registros Proveedores")
    ws_registros.append(COLUMNS_DATA)
    for record in st.session_state.data_records:
        row_data = [record.get(col, '') for col in COLUMNS_DATA]
        ws_registros.append(row_data)
    
    # Hoja de Depósitos
    ws_depositos = wb.create_sheet("Depositos")
    ws_depositos.append(COLUMNS_DEPOSITS)
    for record in st.session_state.deposit_records:
        row_data = [record.get(col, '') for col in COLUMNS_DEPOSITS]
        ws_depositos.append(row_data)
    
    # Hoja de Notas de Débito
    ws_notas = wb.create_sheet("Notas Debito")
    ws_notas.append(COLUMNS_DEBIT_NOTES)
    for record in st.session_state.debit_records:
        row_data = [record.get(col, '') for col in COLUMNS_DEBIT_NOTES]
        ws_notas.append(row_data)
    
    # Hoja de Ventas
    ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                     'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
    ws_ventas = wb.create_sheet("Ventas")
    ws_ventas.append(ventas_columns)
    for record in st.session_state.ventas_records:
        row_data = [record.get(col, '') for col in ventas_columns]
        ws_ventas.append(row_data)
    
    # Hoja de Gastos
    gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
    ws_gastos = wb.create_sheet("Gastos")
    ws_gastos.append(gastos_columns)
    for record in st.session_state.gastos_records:
        row_data = [record.get(col, '') for col in gastos_columns]
        ws_gastos.append(row_data)
    
    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def complete_excel_file(uploaded_file):
    """Completa un archivo Excel existente agregando las hojas faltantes."""
    try:
        # Cargar el archivo existente
        wb = openpyxl.load_workbook(uploaded_file)
        
        # Verificar qué hojas faltan y agregarlas
        hojas_agregadas = []
        
        # Verificar y agregar hoja de Ventas si no existe
        ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                         'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
        
        if "Ventas" not in wb.sheetnames:
            ws_ventas = wb.create_sheet("Ventas")
            ws_ventas.append(ventas_columns)
            hojas_agregadas.append("Ventas")
        
        # Verificar y agregar hoja de Gastos si no existe
        gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
        
        if "Gastos" not in wb.sheetnames:
            ws_gastos = wb.create_sheet("Gastos")
            ws_gastos.append(gastos_columns)
            hojas_agregadas.append("Gastos")
        
        # Si no se agregó ninguna hoja, el archivo ya está completo
        if not hojas_agregadas:
            return uploaded_file.getvalue(), []
        
        # Guardar el archivo completado
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue(), hojas_agregadas
        
    except Exception as e:
        st.error(f"Error completando el archivo Excel: {e}")
        return None, []

def import_from_excel(uploaded_file):
    """Importa datos desde un archivo Excel."""
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        imported_sheets = []
        
        # Lista de posibles nombres para cada hoja (más amplia)
        proveedores_names = ["Registros Proveedores", "proveedores", "Proveedores", "registros_proveedores", 
                            "hoja de proveedores", "registro", "registros", "proveedor"]
        depositos_names = ["Depositos", "depositos", "Depósitos", "Deposito", "depósito", "deposito"]
        notas_names = ["Notas Debito", "notas_debito", "Notas de Débito", "notas de venta", "Notas de Venta", 
                      "notas_venta", "nota", "debito", "débito"]
        ventas_names = ["Ventas", "ventas", "Venta", "venta"]
        gastos_names = ["Gastos", "gastos", "Gasto", "gasto"]
        
        # Función auxiliar para encontrar hoja por nombres posibles
        def find_sheet(possible_names):
            for name in possible_names:
                if name in wb.sheetnames:
                    return name
            return None
        
        # Importar Registros Principales (Proveedores)
        sheet_name = find_sheet(proveedores_names)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            new_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Si la fila no está completamente vacía
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            # Mapear columnas comunes a nuestro formato
                            col_name = headers[i]
                            if col_name in COLUMNS_DATA or any(col.lower() in col_name.lower() for col in COLUMNS_DATA):
                                record[col_name] = str(value) if value is not None else ''
                    if record:  # Solo agregar si tiene datos
                        new_records.append(record)
            
            if new_records:
                st.session_state.data_records = new_records
                imported_sheets.append(f"Proveedores ({sheet_name})")
        
        # Importar Depósitos
        sheet_name = find_sheet(depositos_names)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            new_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            col_name = headers[i]
                            if col_name in COLUMNS_DEPOSITS or any(col.lower() in col_name.lower() for col in COLUMNS_DEPOSITS):
                                record[col_name] = str(value) if value is not None else ''
                    if record:
                        new_records.append(record)
            
            if new_records:
                st.session_state.deposit_records = new_records
                imported_sheets.append(f"Depósitos ({sheet_name})")
        
        # Importar Notas de Débito/Venta
        sheet_name = find_sheet(notas_names)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            new_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            col_name = headers[i]
                            if col_name in COLUMNS_DEBIT_NOTES or any(col.lower() in col_name.lower() for col in COLUMNS_DEBIT_NOTES):
                                record[col_name] = str(value) if value is not None else ''
                    if record:
                        new_records.append(record)
            
            if new_records:
                st.session_state.debit_records = new_records
                imported_sheets.append(f"Notas de Débito ({sheet_name})")
        
        # Importar Ventas (opcional)
        sheet_name = find_sheet(ventas_names)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                             'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
            
            new_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            col_name = headers[i]
                            # Mapear a nombres estándar de ventas
                            if col_name.lower() in [col.lower() for col in ventas_columns]:
                                matching_col = next(col for col in ventas_columns if col.lower() == col_name.lower())
                                record[matching_col] = str(value) if value is not None else ''
                    if record:
                        new_records.append(record)
            
            if new_records:
                st.session_state.ventas_records = new_records
                imported_sheets.append(f"Ventas ({sheet_name})")
        
        # Importar Gastos (opcional)
        sheet_name = find_sheet(gastos_names)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
            
            new_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            col_name = headers[i]
                            # Mapear a nombres estándar de gastos
                            if col_name.lower() in [col.lower() for col in gastos_columns]:
                                matching_col = next(col for col in gastos_columns if col.lower() == col_name.lower())
                                record[matching_col] = str(value) if value is not None else ''
                    if record:
                        new_records.append(record)
            
            if new_records:
                st.session_state.gastos_records = new_records
                imported_sheets.append(f"Gastos ({sheet_name})")
        
        # Guardar en archivos locales y sincronizar con Supabase
        save_all_data()
        recalculate_accumulated_balances()
        sync_to_supabase()
        
        # Mostrar qué hojas se importaron
        if imported_sheets:
            st.success(f"Hojas importadas exitosamente: {', '.join(imported_sheets)}")
        else:
            st.warning("No se encontraron hojas reconocibles para importar. Verifica los nombres de las hojas.")
        
        return len(imported_sheets) > 0
    except Exception as e:
        st.error(f"Error importando desde Excel: {e}")
        return False

def analyze_excel_file(uploaded_file):
    """Analiza un archivo Excel y muestra su estructura."""
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        
        st.write("**📊 Análisis del archivo Excel:**")
        st.write(f"**Hojas encontradas:** {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            st.write(f"**• Hoja: `{sheet_name}`**")
            ws = wb[sheet_name]
            
            # Obtener headers
            headers = []
            if ws.max_row > 0:
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value))
            
            # Contar filas con datos
            data_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data_rows += 1
            
            st.write(f"  - Columnas: {len(headers)}")
            st.write(f"  - Filas con datos: {data_rows}")
            if headers:
                st.write(f"  - Encabezados: {', '.join(headers[:5])}{'...' if len(headers) > 5 else ''}")
            
            # Identificar qué tipo de hoja podría ser
            sheet_type = "Desconocida"
            if any(word in sheet_name.lower() for word in ["proveedor", "registro"]):
                sheet_type = "🏢 Proveedores"
            elif any(word in sheet_name.lower() for word in ["deposito", "depósito"]):
                sheet_type = "💰 Depósitos"
            elif any(word in sheet_name.lower() for word in ["nota", "venta", "débito"]):
                sheet_type = "📋 Notas de Débito/Venta"
            elif any(word in sheet_name.lower() for word in ["venta"]):
                sheet_type = "🛒 Ventas"
            elif any(word in sheet_name.lower() for word in ["gasto"]):
                sheet_type = "💸 Gastos"
            
            st.write(f"  - Tipo identificado: {sheet_type}")
            st.write("---")
            
    except Exception as e:
        st.error(f"Error analizando el archivo: {e}")

def save_all_data():
    """Guarda todos los datos en archivos locales."""
    save_csv_data(st.session_state.data_records, DATA_FILE, COLUMNS_DATA)
    save_csv_data(st.session_state.deposit_records, DEPOSITS_FILE, COLUMNS_DEPOSITS)
    save_csv_data(st.session_state.debit_records, DEBIT_NOTES_FILE, COLUMNS_DEBIT_NOTES)
    
    ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                     'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
    save_csv_data(st.session_state.ventas_records, VENTAS_FILE, ventas_columns)
    
    gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
    save_csv_data(st.session_state.gastos_records, GASTOS_FILE, gastos_columns)

# --- FUNCIONES DE CARGA Y GUARDADO DE DATOS ---

def load_csv_data(file_path, columns):
    """Carga datos desde un archivo CSV."""
    if os.path.exists(file_path):
        try:
            data = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Asegurar que todas las columnas existen
                    record = {}
                    for col in columns:
                        record[col] = row.get(col, '' if col != 'Fecha' else datetime.now().date().strftime('%Y-%m-%d'))
                    data.append(record)
            return data
        except Exception as e:
            st.error(f"Error al cargar {file_path}: {e}")
            return []
    return []

def save_csv_data(data, file_path, columns):
    """Guarda datos en un archivo CSV."""
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for record in data:
                # Convertir fechas a string para CSV
                row = {}
                for col in columns:
                    value = record.get(col, '')
                    if col == 'Fecha' and isinstance(value, date):
                        row[col] = value.strftime('%Y-%m-%d')
                    else:
                        row[col] = value
                writer.writerow(row)
        return True
    except Exception as e:
        st.error(f"Error al guardar {file_path}: {e}")
        return False

# --- FUNCIONES DE LÓGICA DE NEGOCIO ---

def recalculate_accumulated_balances():
    """Recalcula el Saldo Acumulado para todo el sistema."""
    # Obtener datos actuales
    data_records = st.session_state.data_records
    deposit_records = st.session_state.deposit_records
    debit_records = st.session_state.debit_records
    
    # Procesar registros de proveedores
    for i, record in enumerate(data_records):
        if record.get('Proveedor') == 'BALANCE_INICIAL':
            continue
            
        # Convertir valores numéricos
        cantidad = parse_int(record.get('Cantidad', 0))
        peso_salida = parse_float(record.get('Peso Salida (kg)', 0))
        peso_entrada = parse_float(record.get('Peso Entrada (kg)', 0))
        precio_unitario = parse_float(record.get('Precio Unitario ($)', 0))
        
        # Calcular campos derivados
        kilos_restantes = peso_salida - peso_entrada
        libras_restantes = kilos_restantes * LBS_PER_KG
        promedio = libras_restantes / cantidad if cantidad > 0 else 0
        total = libras_restantes * precio_unitario
        
        # Actualizar registro
        data_records[i]['Kilos Restantes'] = kilos_restantes
        data_records[i]['Libras Restantes'] = libras_restantes
        data_records[i]['Promedio'] = promedio
        data_records[i]['Total ($)'] = total
    
    # Calcular montos de depósito por fecha y empresa
    deposit_by_date_empresa = {}
    for dep in deposit_records:
        fecha = dep.get('Fecha', '')
        empresa = dep.get('Empresa', '')
        monto = parse_float(dep.get('Monto', 0))
        key = f"{fecha}_{empresa}"
        deposit_by_date_empresa[key] = deposit_by_date_empresa.get(key, 0) + monto
    
    # Asignar montos de depósito a registros
    for i, record in enumerate(data_records):
        if record.get('Proveedor') == 'BALANCE_INICIAL':
            continue
        fecha = record.get('Fecha', '')
        proveedor = record.get('Proveedor', '')
        key = f"{fecha}_{proveedor}"
        data_records[i]['Monto Deposito'] = deposit_by_date_empresa.get(key, 0)
    
    # Calcular saldos diarios por fecha
    daily_balances = {}
    for record in data_records:
        if record.get('Proveedor') == 'BALANCE_INICIAL':
            continue
        fecha = record.get('Fecha', '')
        monto_deposito = parse_float(record.get('Monto Deposito', 0))
        total = parse_float(record.get('Total ($)', 0))
        saldo_diario = monto_deposito - total
        
        if fecha not in daily_balances:
            daily_balances[fecha] = 0
        daily_balances[fecha] += saldo_diario
    
    # Incorporar notas de débito
    for note in debit_records:
        fecha = note.get('Fecha', '')
        descuento_real = parse_float(note.get('Descuento real', 0))
        if fecha in daily_balances:
            daily_balances[fecha] += descuento_real
    
    # Calcular saldos acumulados
    sorted_dates = sorted(daily_balances.keys())
    accumulated_balance = INITIAL_ACCUMULATED_BALANCE
    accumulated_by_date = {}
    
    for fecha in sorted_dates:
        accumulated_balance += daily_balances[fecha]
        accumulated_by_date[fecha] = accumulated_balance
    
    # Asignar saldos a registros
    for i, record in enumerate(data_records):
        if record.get('Proveedor') == 'BALANCE_INICIAL':
            data_records[i]['Saldo Acumulado'] = INITIAL_ACCUMULATED_BALANCE
            data_records[i]['Saldo diario'] = 0
            continue
        
        fecha = record.get('Fecha', '')
        data_records[i]['Saldo diario'] = daily_balances.get(fecha, 0)
        data_records[i]['Saldo Acumulado'] = accumulated_by_date.get(fecha, INITIAL_ACCUMULATED_BALANCE)

def get_next_n():
    """Genera el siguiente número 'N' para un registro."""
    max_n = 0
    for record in st.session_state.data_records:
        if record.get('Proveedor') != 'BALANCE_INICIAL':
            try:
                n_val = int(record.get('N', '0'))
                max_n = max(max_n, n_val)
            except:
                pass
    return f"{max_n + 1:02d}"

# --- FUNCIONES DE AGREGADO DE REGISTROS ---

def add_deposit_record(fecha_d, empresa, agencia, monto):
    """Agrega un nuevo registro de depósito."""
    max_n = 0
    for record in st.session_state.deposit_records:
        try:
            n_val = int(record.get('N', '0'))
            max_n = max(max_n, n_val)
        except:
            pass
    
    numero = f"{max_n + 1:02d}"
    documento = "Deposito" if "Cajero" in agencia else "Transferencia"
    
    nuevo_registro = {
        "Fecha": fecha_d.strftime('%Y-%m-%d'),
        "Empresa": empresa,
        "Agencia": agencia,
        "Monto": str(monto),
        "Documento": documento,
        "N": numero
    }
    
    st.session_state.deposit_records.append(nuevo_registro)
    if save_csv_data(st.session_state.deposit_records, DEPOSITS_FILE, COLUMNS_DEPOSITS):
        st.session_state.deposit_added = True
        sync_to_supabase()  # Sincronizar con Supabase
        st.success("Depósito agregado exitosamente. Recalculando saldos...")
    else:
        st.error("Error al guardar el depósito.")

def add_supplier_record(fecha, proveedor, cantidad, peso_salida, peso_entrada, tipo_documento, gavetas, precio_unitario):
    """Agrega un nuevo registro de proveedor."""
    # Validaciones
    if not all(val >= 0 for val in [cantidad, peso_salida, peso_entrada, precio_unitario, gavetas]):
        st.error("Los valores numéricos no pueden ser negativos.")
        return False
    if cantidad == 0 and peso_salida == 0 and peso_entrada == 0:
        st.error("Por favor, ingresa una Cantidad y/o Pesos válidos.")
        return False
    if peso_entrada > peso_salida:
        st.error("El Peso Entrada no puede ser mayor que el Peso Salida.")
        return False

    # Cálculos
    kilos_restantes = peso_salida - peso_entrada
    libras_restantes = kilos_restantes * LBS_PER_KG
    promedio = libras_restantes / cantidad if cantidad > 0 else 0
    total = libras_restantes * precio_unitario

    enumeracion = get_next_n()

    nueva_fila = {
        "N": enumeracion,
        "Fecha": fecha.strftime('%Y-%m-%d'),
        "Proveedor": proveedor,
        "Producto": PRODUCT_NAME,
        "Cantidad": str(cantidad),
        "Peso Salida (kg)": str(peso_salida),
        "Peso Entrada (kg)": str(peso_entrada),
        "Tipo Documento": tipo_documento,
        "Cantidad de gavetas": str(gavetas),
        "Precio Unitario ($)": str(precio_unitario),
        "Promedio": str(promedio),
        "Kilos Restantes": str(kilos_restantes),
        "Libras Restantes": str(libras_restantes),
        "Total ($)": str(total),
        "Monto Deposito": "0",
        "Saldo diario": "0",
        "Saldo Acumulado": str(INITIAL_ACCUMULATED_BALANCE)
    }

    st.session_state.data_records.append(nueva_fila)
    
    if save_csv_data(st.session_state.data_records, DATA_FILE, COLUMNS_DATA):
        st.session_state.record_added = True
        sync_to_supabase()  # Sincronizar con Supabase
        st.success("Registro agregado correctamente. Recalculando saldos...")
        return True
    else:
        st.error("Error al guardar el registro.")
        return False

def add_debit_note(fecha_nota, descuento, descuento_real):
    """Agrega una nueva nota de débito."""
    # Calcular libras de la fecha
    libras_calculadas = 0
    for record in st.session_state.data_records:
        if (record.get('Fecha') == fecha_nota.strftime('%Y-%m-%d') and 
            record.get('Proveedor') != 'BALANCE_INICIAL'):
            libras_calculadas += parse_float(record.get('Libras Restantes', 0))
    
    descuento_posible = libras_calculadas * descuento
    
    nueva_nota = {
        "Fecha": fecha_nota.strftime('%Y-%m-%d'),
        "Libras calculadas": str(libras_calculadas),
        "Descuento": str(descuento),
        "Descuento posible": str(descuento_posible),
        "Descuento real": str(descuento_real)
    }
    
    st.session_state.debit_records.append(nueva_nota)
    if save_csv_data(st.session_state.debit_records, DEBIT_NOTES_FILE, COLUMNS_DEBIT_NOTES):
        st.session_state.debit_note_added = True
        sync_to_supabase()  # Sincronizar con Supabase
        st.success("Nota de débito agregada correctamente. Recalculando saldos...")
    else:
        st.error("Error al guardar la nota de débito.")

# --- FUNCIONES DEL SISTEMA SECUNDARIO ---

def guardar_venta(venta_data):
    """Guarda una nueva venta."""
    st.session_state.ventas_records.append(venta_data)
    ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                     'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
    if save_csv_data(st.session_state.ventas_records, VENTAS_FILE, ventas_columns):
        sync_to_supabase()  # Sincronizar con Supabase
        st.success("Venta registrada exitosamente")
    else:
        st.error("Error al guardar la venta")

def guardar_gasto(gasto_data):
    """Guarda un nuevo gasto."""
    st.session_state.gastos_records.append(gasto_data)
    gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
    if save_csv_data(st.session_state.gastos_records, GASTOS_FILE, gastos_columns):
        sync_to_supabase()  # Sincronizar con Supabase
        st.success("Gasto registrado exitosamente")
    else:
        st.error("Error al guardar el gasto")

def analizar_alertas_clientes():
    """Analiza las ventas para identificar clientes con alertas."""
    alertas = []
    saldos_por_cliente = {}
    
    for venta in st.session_state.ventas_records:
        cliente = venta.get('cliente', '')
        saldo = parse_float(venta.get('saldo', 0))
        if saldo > 0:
            saldos_por_cliente[cliente] = saldos_por_cliente.get(cliente, 0) + saldo
    
    for cliente, saldo_total in saldos_por_cliente.items():
        if saldo_total > 50:  # Umbral de alerta
            alertas.append(f"⚠️ {cliente}: Saldo pendiente de {formatear_moneda(saldo_total)}")
    
    return alertas

# --- INICIALIZACIÓN DEL ESTADO ---

def initialize_session_state():
    """Inicializa todos los datos en st.session_state."""
    # Inicializar registros principales
    if "data_records" not in st.session_state:
        st.session_state.data_records = load_csv_data(DATA_FILE, COLUMNS_DATA)
        
        # Asegurar que existe el balance inicial
        balance_exists = any(record.get('Proveedor') == 'BALANCE_INICIAL' 
                           for record in st.session_state.data_records)
        
        if not balance_exists:
            balance_record = {col: '' for col in COLUMNS_DATA}
            balance_record.update({
                "Fecha": "1900-01-01",
                "Proveedor": "BALANCE_INICIAL",
                "Saldo diario": "0",
                "Saldo Acumulado": str(INITIAL_ACCUMULATED_BALANCE),
                "Monto Deposito": "0",
                "Total ($)": "0",
                "N": "00"
            })
            st.session_state.data_records.insert(0, balance_record)

    if "deposit_records" not in st.session_state:
        st.session_state.deposit_records = load_csv_data(DEPOSITS_FILE, COLUMNS_DEPOSITS)

    if "debit_records" not in st.session_state:
        st.session_state.debit_records = load_csv_data(DEBIT_NOTES_FILE, COLUMNS_DEBIT_NOTES)

    # Inicializar registros de ventas y gastos
    if "ventas_records" not in st.session_state:
        ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                         'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
        st.session_state.ventas_records = load_csv_data(VENTAS_FILE, ventas_columns)

    if "gastos_records" not in st.session_state:
        gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
        st.session_state.gastos_records = load_csv_data(GASTOS_FILE, gastos_columns)

    # Intentar cargar desde Supabase al inicio
    if "supabase_loaded" not in st.session_state:
        if load_from_supabase():
            st.session_state.supabase_loaded = True

    # Recalcular saldos
    recalculate_accumulated_balances()
    
    # Inicializar flags
    for flag in ["deposit_added", "record_added", "debit_note_added", "supabase_loaded"]:
        if flag not in st.session_state:
            st.session_state[flag] = False

# --- INTERFAZ DE USUARIO ---

def render_deposit_form():
    """Formulario de registro de depósitos."""
    st.sidebar.subheader("📋 Registrar Depósito")
    
    with st.sidebar.form(key="form_deposito", clear_on_submit=True):
        fecha_deposito = st.date_input("Fecha del depósito", value=datetime.now().date())
        empresa_deposito = st.selectbox("Empresa (Proveedor)", options=PROVEEDORES)
        agencia_deposito = st.selectbox("Agencia", options=AGENCIAS)
        monto_deposito = st.number_input("Monto ($)", min_value=0.01, step=0.01)
        
        if st.form_submit_button("💾 Agregar Depósito"):
            add_deposit_record(fecha_deposito, empresa_deposito, agencia_deposito, monto_deposito)

def render_supplier_form():
    """Formulario de registro de proveedores."""
    st.subheader("📝 Registrar Nueva Transacción de Proveedor")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        fecha_registro = st.date_input("Fecha", value=datetime.now().date())
        proveedor = st.selectbox("Proveedor", options=PROVEEDORES)
        cantidad = st.number_input("Cantidad", min_value=0, step=1)
    
    with col2:
        peso_salida = st.number_input("Peso Salida (kg)", min_value=0.0, step=0.01)
        peso_entrada = st.number_input("Peso Entrada (kg)", min_value=0.0, step=0.01)
        tipo_documento = st.selectbox("Tipo de Documento", options=TIPOS_DOCUMENTO)
    
    with col3:
        gavetas = st.number_input("Cantidad de gavetas", min_value=0, step=1)
        precio_unitario = st.number_input("Precio Unitario ($)", min_value=0.01, step=0.01)
    
    if st.button("💾 Agregar Registro"):
        add_supplier_record(fecha_registro, proveedor, cantidad, peso_salida, 
                          peso_entrada, tipo_documento, gavetas, precio_unitario)

def render_debit_form():
    """Formulario de notas de débito."""
    st.subheader("📄 Agregar Nota de Débito")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        fecha_nota = st.date_input("Fecha de la nota", value=datetime.now().date())
    with col2:
        descuento_porcentaje = st.number_input("Descuento (%)", min_value=0.0, max_value=1.0, step=0.01)
    with col3:
        descuento_real_input = st.number_input("Descuento Real ($)", min_value=0.0, step=0.01)
    
    if st.button("💾 Agregar Nota de Débito"):
        add_debit_note(fecha_nota, descuento_porcentaje, descuento_real_input)

def render_data_tables():
    """Renderiza las tablas de datos."""
    st.subheader("📊 Tablas de Datos")
    
    tab1, tab2, tab3 = st.tabs(["Registros Principales", "Depósitos", "Notas de Débito"])
    
    with tab1:
        st.write("**Registros de Proveedores**")
        if st.session_state.data_records:
            # Mostrar tabla como HTML para evitar problemas con pyarrow
            html_table = "<table style='width:100%; border-collapse: collapse;'>"
            html_table += "<tr style='background-color: #f0f2f6;'>"
            for col in COLUMNS_DATA:
                html_table += f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{col}</th>"
            html_table += "</tr>"
            
            for record in st.session_state.data_records:
                html_table += "<tr>"
                for col in COLUMNS_DATA:
                    value = record.get(col, '')
                    if col in ["Precio Unitario ($)", "Total ($)", "Monto Deposito", "Saldo diario", "Saldo Acumulado"]:
                        value = formatear_moneda(value)
                    html_table += f"<td style='border: 1px solid #ddd; padding: 8px;'>{value}</td>"
                html_table += "</tr>"
            html_table += "</table>"
            
            st.markdown(html_table, unsafe_allow_html=True)
            
            # Mostrar resumen de registros
            total_registros = len([r for r in st.session_state.data_records if r.get('Proveedor') != 'BALANCE_INICIAL'])
            st.write(f"**Total de registros:** {total_registros}")
        else:
            st.info("No hay registros disponibles.")
    
    with tab2:
        st.write("**Depósitos**")
        if st.session_state.deposit_records:
            html_table = "<table style='width:100%; border-collapse: collapse;'>"
            html_table += "<tr style='background-color: #f0f2f6;'>"
            for col in COLUMNS_DEPOSITS:
                html_table += f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{col}</th>"
            html_table += "</tr>"
            
            for record in st.session_state.deposit_records:
                html_table += "<tr>"
                for col in COLUMNS_DEPOSITS:
                    value = record.get(col, '')
                    if col == "Monto":
                        value = formatear_moneda(value)
                    html_table += f"<td style='border: 1px solid #ddd; padding: 8px;'>{value}</td>"
                html_table += "</tr>"
            html_table += "</table>"
            
            st.markdown(html_table, unsafe_allow_html=True)
            
            total_depositos = sum(parse_float(d.get('Monto', 0)) for d in st.session_state.deposit_records)
            st.write(f"**Total depósitos:** {formatear_moneda(total_depositos)}")
        else:
            st.info("No hay depósitos registrados.")
    
    with tab3:
        st.write("**Notas de Débito**")
        if st.session_state.debit_records:
            html_table = "<table style='width:100%; border-collapse: collapse;'>"
            html_table += "<tr style='background-color: #f0f2f6;'>"
            for col in COLUMNS_DEBIT_NOTES:
                html_table += f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{col}</th>"
            html_table += "</tr>"
            
            for record in st.session_state.debit_records:
                html_table += "<tr>"
                for col in COLUMNS_DEBIT_NOTES:
                    value = record.get(col, '')
                    if col in ["Descuento posible", "Descuento real"]:
                        value = formatear_moneda(value)
                    html_table += f"<td style='border: 1px solid #ddd; padding: 8px;'>{value}</td>"
                html_table += "</tr>"
            html_table += "</table>"
            
            st.markdown(html_table, unsafe_allow_html=True)
            
            total_notas = sum(parse_float(n.get('Descuento real', 0)) for n in st.session_state.debit_records)
            st.write(f"**Total notas de débito:** {formatear_moneda(total_notas)}")
        else:
            st.info("No hay notas de débito registradas.")
    
    # Sección de descarga e importación
    st.subheader("📥📤 Gestión de Archivos Excel")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Descargar Datos**")
        if st.button("📊 Descargar Todo en Excel"):
            excel_data = create_excel_export()
            st.download_button(
                label="💾 Descargar archivo Excel",
                data=excel_data,
                file_name=f"sistema_gestion_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        st.write("**Importar Datos**")
        uploaded_file = st.file_uploader(
            "Selecciona un archivo Excel para importar",
            type=['xlsx'],
            help="El archivo puede tener diferentes nombres de hojas - el sistema se adaptará automáticamente"
        )
        
        if uploaded_file is not None:
            # Botón para diagnosticar el archivo
            if st.button("🔍 Analizar Archivo Excel"):
                analyze_excel_file(uploaded_file)
            
            # Botón para completar el archivo con hojas faltantes
            if st.button("📋 Completar y Descargar Archivo"):
                completed_file, added_sheets = complete_excel_file(uploaded_file)
                if completed_file:
                    if added_sheets:
                        st.success(f"Archivo completado. Hojas agregadas: {', '.join(added_sheets)}")
                    else:
                        st.info("El archivo ya tenía todas las hojas necesarias")
                    
                    # Botón de descarga para el archivo completado
                    st.download_button(
                        label="💾 Descargar Archivo Completado",
                        data=completed_file,
                        file_name=f"archivo_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            # Botón para importar
            if st.button("📋 Importar desde Excel"):
                if import_from_excel(uploaded_file):
                    st.rerun()
                else:
                    st.error("Error al importar datos desde Excel")
    
    # Sección de sincronización con Supabase
    st.subheader("🔄 Sincronización con Supabase")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("☁️ Sincronizar con Supabase"):
            if sync_to_supabase():
                st.success("Datos sincronizados correctamente con Supabase")
            else:
                st.error("Error al sincronizar con Supabase")
    
    with col2:
        if st.button("📥 Cargar desde Supabase"):
            if load_from_supabase():
                recalculate_accumulated_balances()
                st.success("Datos cargados correctamente desde Supabase")
                st.rerun()
            else:
                st.error("Error al cargar datos desde Supabase")

def render_ventas_section():
    """Sección de gestión de ventas."""
    st.header("🛒 Gestión de Ventas")
    
    st.subheader("📝 Registrar Nueva Venta")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        fecha_venta = st.date_input("Fecha", value=datetime.now().date(), key="fecha_venta")
        cliente_venta = st.selectbox("Cliente", options=CLIENTES, key="cliente_venta")
    
    with col2:
        tipo_ave = st.selectbox("Tipo de Ave", options=TIPOS_AVE, key="tipo_ave")
        cantidad_venta = st.number_input("Cantidad", min_value=1, step=1, key="cantidad_venta")
    
    with col3:
        libras_venta = st.number_input("Libras", min_value=0.1, step=0.1, key="libras_venta")
        descuento_venta = st.number_input("Descuento (libras)", min_value=0.0, step=0.1, key="descuento_venta")
    
    with col4:
        precio_venta = st.number_input("Precio por libra ($)", min_value=0.01, step=0.01, key="precio_venta")
        pago_cliente = st.number_input("Pago del Cliente ($)", min_value=0.0, step=0.01, key="pago_cliente")
    
    # Cálculos automáticos
    libras_netas = libras_venta - descuento_venta
    total_a_cobrar = libras_netas * precio_venta
    saldo_pendiente = total_a_cobrar - pago_cliente
    
    col1_calc, col2_calc, col3_calc = st.columns(3)
    with col1_calc:
        st.metric("Libras Netas", f"{libras_netas:.2f} lbs")
    with col2_calc:
        st.metric("Total a Cobrar", formatear_moneda(total_a_cobrar))
    with col3_calc:
        st.metric("Saldo Pendiente", formatear_moneda(saldo_pendiente))
    
    if st.button("💾 Registrar Venta"):
        venta_data = {
            'fecha': fecha_venta.strftime('%Y-%m-%d'),
            'cliente': cliente_venta,
            'tipo': tipo_ave,
            'cantidad': str(cantidad_venta),
            'libras': str(libras_venta),
            'descuento': str(descuento_venta),
            'libras_netas': str(libras_netas),
            'precio': str(precio_venta),
            'total_a_cobrar': str(total_a_cobrar),
            'pago_cliente': str(pago_cliente),
            'saldo': str(saldo_pendiente)
        }
        guardar_venta(venta_data)
    
    # Mostrar alertas
    alertas = analizar_alertas_clientes()
    if alertas:
        st.warning("**Alertas de Clientes:**")
        for alerta in alertas:
            st.write(alerta)
    
    # Tabla de ventas
    st.subheader("📊 Historial de Ventas")
    if st.session_state.ventas_records:
        ventas_columns = ['fecha', 'cliente', 'tipo', 'cantidad', 'libras', 'descuento',
                         'libras_netas', 'precio', 'total_a_cobrar', 'pago_cliente', 'saldo']
        
        html_table = "<table style='width:100%; border-collapse: collapse;'>"
        html_table += "<tr style='background-color: #f0f2f6;'>"
        for col in ventas_columns:
            html_table += f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{col.replace('_', ' ').title()}</th>"
        html_table += "</tr>"
        
        for venta in st.session_state.ventas_records:
            html_table += "<tr>"
            for col in ventas_columns:
                value = venta.get(col, '')
                if col in ['total_a_cobrar', 'pago_cliente', 'saldo']:
                    value = formatear_moneda(value)
                html_table += f"<td style='border: 1px solid #ddd; padding: 8px;'>{value}</td>"
            html_table += "</tr>"
        html_table += "</table>"
        
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Resumen
        total_ventas = sum(parse_float(v.get('total_a_cobrar', 0)) for v in st.session_state.ventas_records)
        total_cobrado = sum(parse_float(v.get('pago_cliente', 0)) for v in st.session_state.ventas_records)
        total_pendiente = sum(parse_float(v.get('saldo', 0)) for v in st.session_state.ventas_records)
        
        col1_res, col2_res, col3_res = st.columns(3)
        with col1_res:
            st.metric("Total Ventas", formatear_moneda(total_ventas))
        with col2_res:
            st.metric("Total Cobrado", formatear_moneda(total_cobrado))
        with col3_res:
            st.metric("Total Pendiente", formatear_moneda(total_pendiente))
    else:
        st.info("No hay ventas registradas.")

def render_gastos_section():
    """Sección de gestión de gastos."""
    st.header("💰 Gestión de Gastos")
    
    st.subheader("📝 Registrar Nuevo Gasto")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        fecha_gasto = st.date_input("Fecha", value=datetime.now().date(), key="fecha_gasto")
    with col2:
        categoria_gasto = st.selectbox("Categoría", options=CATEGORIAS_GASTO, key="categoria_gasto")
    with col3:
        cantidad_gasto = st.number_input("Cantidad ($)", min_value=0.01, step=0.01, key="cantidad_gasto")
    
    descripcion_gasto = st.text_area("Descripción", placeholder="Describe el gasto...", key="descripcion_gasto")
    
    if st.button("💾 Registrar Gasto"):
        gasto_data = {
            'fecha': fecha_gasto.strftime('%Y-%m-%d'),
            'categoria': categoria_gasto,
            'descripcion': descripcion_gasto,
            'cantidad': str(cantidad_gasto)
        }
        guardar_gasto(gasto_data)
    
    # Tabla de gastos
    st.subheader("📊 Historial de Gastos")
    if st.session_state.gastos_records:
        gastos_columns = ['fecha', 'categoria', 'descripcion', 'cantidad']
        
        html_table = "<table style='width:100%; border-collapse: collapse;'>"
        html_table += "<tr style='background-color: #f0f2f6;'>"
        for col in gastos_columns:
            html_table += f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{col.replace('_', ' ').title()}</th>"
        html_table += "</tr>"
        
        for gasto in st.session_state.gastos_records:
            html_table += "<tr>"
            for col in gastos_columns:
                value = gasto.get(col, '')
                if col == 'cantidad':
                    value = formatear_moneda(value)
                html_table += f"<td style='border: 1px solid #ddd; padding: 8px;'>{value}</td>"
            html_table += "</tr>"
        html_table += "</table>"
        
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Resumen
        total_gastos = sum(parse_float(g.get('cantidad', 0)) for g in st.session_state.gastos_records)
        st.metric("Total Gastos", formatear_moneda(total_gastos))
    else:
        st.info("No hay gastos registrados.")

# --- APLICACIÓN PRINCIPAL ---

def main():
    """Función principal de la aplicación."""
    initialize_session_state()
    
    st.title("🏢 Sistema Integral de Gestión de Proveedores y Ventas")
    st.markdown("---")
    
    # Sidebar
    with st.sidebar:
        st.title("🔧 Panel de Control")
        st.markdown("---")
        
        # Mostrar saldo actual
        if st.session_state.data_records:
            saldo_actual = INITIAL_ACCUMULATED_BALANCE
            for record in st.session_state.data_records:
                if record.get('Proveedor') != 'BALANCE_INICIAL':
                    saldo_actual = parse_float(record.get('Saldo Acumulado', INITIAL_ACCUMULATED_BALANCE))
            st.metric("💰 Saldo Actual", formatear_moneda(saldo_actual))
        
        st.markdown("---")
        render_deposit_form()
        st.markdown("---")
        
        # Información del sistema
        st.info(f"""
        **Registros Totales:**
        - Proveedores: {len([r for r in st.session_state.data_records if r.get('Proveedor') != 'BALANCE_INICIAL'])}
        - Depósitos: {len(st.session_state.deposit_records)}
        - Notas Débito: {len(st.session_state.debit_records)}
        - Ventas: {len(st.session_state.ventas_records)}
        - Gastos: {len(st.session_state.gastos_records)}
        """)
    
    # Navegación principal
    main_tab = st.selectbox(
        "📂 Selecciona una sección:",
        ["🏭 Gestión de Proveedores", "🛒 Gestión de Ventas", "💰 Gestión de Gastos", "📊 Ver Datos"]
    )
    
    if main_tab == "🏭 Gestión de Proveedores":
        st.header("🏭 Gestión de Proveedores")
        provider_tabs = st.tabs(["Nuevo Registro", "Nota de Débito", "Ver Datos"])
        
        with provider_tabs[0]:
            render_supplier_form()
        with provider_tabs[1]:
            render_debit_form()
        with provider_tabs[2]:
            render_data_tables()
    
    elif main_tab == "🛒 Gestión de Ventas":
        render_ventas_section()
    
    elif main_tab == "💰 Gestión de Gastos":
        render_gastos_section()
    
    elif main_tab == "📊 Ver Datos":
        render_data_tables()
    
    # Recalcular saldos si hay cambios
    if any([st.session_state.get(flag, False) for flag in ["deposit_added", "record_added", "debit_note_added"]]):
        recalculate_accumulated_balances()
        for flag in ["deposit_added", "record_added", "debit_note_added"]:
            st.session_state[flag] = False
        st.rerun()

if __name__ == "__main__":
    main()
